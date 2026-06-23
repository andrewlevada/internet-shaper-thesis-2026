#!/usr/bin/env python3
"""Convert a request-prompts CSV to an evaluation XLSX with embedded screenshots.

Input:  ../our-3-request-prompts.csv   (columns: number, request-prompt)
Output: ../our-3-request-prompts.xlsx  (columns: number, verdict, request-prompt, screenshot)

Screenshots are loaded from ../seed-samples/our-3/{number}/screenshot.png.
The verdict column is intentionally left empty for manual review.
"""

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent

INPUT_CSV = DATA_DIR / "our-3-request-prompts.csv"
OUTPUT_XLSX = DATA_DIR / "our-3-request-prompts.xlsx"
SCREENSHOTS_DIR = DATA_DIR / "seed-samples" / "our-3"

# Column layout (1-indexed)
COL_NUMBER = 1      # A
COL_VERDICT = 2     # B
COL_PROMPT = 3      # C
COL_SCREENSHOT = 4  # D

# Dimensions (in Excel units)
ROW_HEIGHT_HEADER_TITLE = 27.65
ROW_HEIGHT_HEADER_COLS = 15.0
ROW_HEIGHT_DATA = 174.75

COL_WIDTHS = {
    get_column_letter(COL_NUMBER): 7.67,
    get_column_letter(COL_VERDICT): 9.85,
    get_column_letter(COL_PROMPT): 60.0,
    get_column_letter(COL_SCREENSHOT): 93.5,
}

# Approximate pixel height for the screenshot cell (used for image scaling)
# Excel row height unit ≈ 0.75 pt; 1 pt ≈ 1.333 px
ROW_HEIGHT_PX = int(ROW_HEIGHT_DATA * 0.75 * 1.333)

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TITLE_FONT = Font(bold=True, size=12)
COL_HEADER_FONT = Font(bold=True)


def read_csv(path: Path) -> list[dict]:
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def scale_image_to_height(img_path: Path, target_height_px: int) -> XLImage:
    """Load image and scale it proportionally to fit within target_height_px."""
    img = XLImage(str(img_path))
    # openpyxl stores raw pixel dimensions in img.width / img.height
    if img.height and img.height > 0:
        scale = target_height_px / img.height
        img.width = int(img.width * scale)
        img.height = target_height_px
    return img


def build_xlsx(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Request Prompts"

    # --- Column widths ---
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # --- Row 1: title ---
    ws.row_dimensions[1].height = ROW_HEIGHT_HEADER_TITLE
    title_cell = ws.cell(row=1, column=COL_NUMBER, value="our-3-request-prompts")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- Row 2: column headers ---
    ws.row_dimensions[2].height = ROW_HEIGHT_HEADER_COLS
    headers = {
        COL_NUMBER: "number",
        COL_VERDICT: "verdict",
        COL_PROMPT: "request-prompt",
        COL_SCREENSHOT: "screenshot",
    }
    for col, label in headers.items():
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = COL_HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Data rows ---
    for i, row in enumerate(rows):
        excel_row = i + 3  # rows 3+
        ws.row_dimensions[excel_row].height = ROW_HEIGHT_DATA

        number = row["number"]
        prompt = row["request-prompt"]

        ws.cell(row=excel_row, column=COL_NUMBER, value=number).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        ws.cell(row=excel_row, column=COL_VERDICT, value=None)  # empty for review
        prompt_cell = ws.cell(row=excel_row, column=COL_PROMPT, value=prompt)
        prompt_cell.alignment = Alignment(wrap_text=True, vertical="top")

        screenshot_path = SCREENSHOTS_DIR / number / "original" / "screenshot.png"
        if screenshot_path.exists():
            img = scale_image_to_height(screenshot_path, ROW_HEIGHT_PX)
            # Anchor at screenshot column (0-indexed col = COL_SCREENSHOT - 1)
            col_letter = get_column_letter(COL_SCREENSHOT)
            img.anchor = f"{col_letter}{excel_row}"
            ws.add_image(img)
        else:
            ws.cell(row=excel_row, column=COL_SCREENSHOT, value="(screenshot missing)")

    wb.save(OUTPUT_XLSX)
    print(f"Saved: {OUTPUT_XLSX}")


def main() -> None:
    rows = read_csv(INPUT_CSV)
    print(f"Read {len(rows)} rows from {INPUT_CSV.name}")

    missing = [
        r["number"]
        for r in rows
        if not (SCREENSHOTS_DIR / r["number"] / "original" / "screenshot.png").exists()
    ]
    if missing:
        print(f"Warning: {len(missing)} screenshots missing: {missing[:5]}{'...' if len(missing) > 5 else ''}")

    build_xlsx(rows)
    print(f"Done. {len(rows) - len(missing)}/{len(rows)} screenshots embedded.")


if __name__ == "__main__":
    main()
