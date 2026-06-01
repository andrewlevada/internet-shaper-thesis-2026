#!/usr/bin/env python3
"""Filter out (clear out) selected seed-samples based on a verdict column in an XLSX file.

For the selected sample folder under seed-samples/, processes every entry listed in the XLSX:
  1. If a seed-sample's task.json already has "protected_at", it is skipped and left completely unchanged.
  2. If the XLSX row has a filled "verdict" column, the seed-sample folder is deleted.
  3. Otherwise (the "remaining" sample), a "protected_at" timestamp is written into task.json.

Usage:
  python3 filter-out-request-prompts.py --xlsx our-3-request-prompts.xlsx --sample our-3
  python3 filter-out-request-prompts.py --xlsx our-3-request-prompts.xlsx --sample our-3 --dry-run
"""

from __future__ import annotations

import argparse
import json
import shutil
import time
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR.parent
SEED_SAMPLES_DIR = DATA_DIR / "seed-samples"
LOGS_DIR = SCRIPT_DIR / "logs"

SAMPLE_WIDTH = 3


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Filter out request prompts based on a filled verdict column in an XLSX file.",
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        help="Path or name of the XLSX file (e.g. our-3-request-prompts.xlsx).",
    )
    parser.add_argument(
        "--sample",
        required=True,
        help="The seed-samples subfolder name to process (e.g. our-3).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print planned actions without deleting folders or updating task.json files.",
    )
    return parser.parse_args()


def load_task_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def write_task_json(path: Path, task: dict) -> None:
    path.write_text(
        json.dumps(task, indent="\t", ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def format_sample_number(val: any) -> str | None:
    """Format row's 'number' column to a 3-digit padded string (e.g. '001')."""
    if val is None:
        return None
    try:
        # Excel might parse '001' as integer 1
        return f"{int(val):0{SAMPLE_WIDTH}d}"
    except (ValueError, TypeError):
        s = str(val).strip()
        if not s:
            return None
        # Try to format any numeric part or just pad with zeros if numeric
        if s.isdigit():
            return f"{int(s):0{SAMPLE_WIDTH}d}"
        return s


def main() -> None:
    args = parse_args()
    timestamp = datetime.now(timezone.utc).isoformat()
    file_ts = timestamp.replace(":", "-")
    log_path = LOGS_DIR / f"{file_ts}-filter-out-request-prompts.log"
    log_lines: list[str] = []

    # Resolve xlsx path
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        if not xlsx_path.exists():
            xlsx_path = DATA_DIR / args.xlsx
    
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    # Resolve seed-samples directory
    fold_dir = SEED_SAMPLES_DIR / args.sample
    if not fold_dir.is_dir():
        raise FileNotFoundError(f"Seed samples folder not found: {fold_dir}")

    log_lines.extend(
        [
            "Filter out request prompts script run",
            "-" * 48,
            f"timestamp (UTC): {timestamp}",
            f"script: {Path(__file__).name}",
            f"xlsx file: {xlsx_path}",
            f"sample fold: {fold_dir}",
            f"dry run: {args.dry_run}",
            "",
        ]
    )

    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    deleted_count = 0
    protected_new_count = 0
    already_protected_count = 0
    missing_count = 0

    log_lines.append(f"Processing rows from active sheet '{ws.title}'...")

    # Excel rows: Row 1 is title, Row 2 is headers, data starts at Row 3.
    # Columns: 1=number, 2=verdict
    for row_idx in range(3, ws.max_row + 1):
        num_val = ws.cell(row=row_idx, column=1).value
        verdict_val = ws.cell(row=row_idx, column=2).value

        number_str = format_sample_number(num_val)
        if not number_str:
            continue

        sample_dir = fold_dir / number_str
        task_path = sample_dir / "task.json"

        # Check if the folder exists at all
        if not sample_dir.is_dir():
            msg = f"skip {number_str}: folder does not exist at {sample_dir}"
            log_lines.append(msg)
            print(msg)
            missing_count += 1
            continue

        # Check task.json for existing protection stamp
        has_stamp = False
        if task_path.is_file():
            try:
                task = load_task_json(task_path)
                if "protected_at" in task:
                    has_stamp = True
            except Exception as exc:
                msg = f"WARNING {number_str}: failed to load task.json ({exc})"
                log_lines.append(msg)
                print(msg)

        if has_stamp:
            msg = f"skip {number_str}: already protected by stamp in task.json (left unchanged)"
            log_lines.append(msg)
            print(msg)
            already_protected_count += 1
            continue

        # Check if verdict column is filled
        is_verdict_filled = verdict_val is not None and str(verdict_val).strip() != ""

        if is_verdict_filled:
            # Verdict is filled -> clear out / delete
            if args.dry_run:
                msg = f"would clear out {number_str} (verdict: {verdict_val!r})"
                log_lines.append(msg)
                print(msg)
            else:
                shutil.rmtree(sample_dir)
                msg = f"OK cleared out {number_str} (deleted folder, verdict: {verdict_val!r})"
                log_lines.append(msg)
                print(msg)
            deleted_count += 1
        else:
            # Verdict is empty -> remaining sample. Stamp it!
            if task_path.is_file():
                protected_at = int(time.time() * 1000)
                if args.dry_run:
                    msg = f"would protect {number_str} (leaving protection timestamp protected_at={protected_at})"
                    log_lines.append(msg)
                    print(msg)
                else:
                    try:
                        task = load_task_json(task_path)
                        task["protected_at"] = protected_at
                        write_task_json(task_path, task)
                        msg = f"OK protected {number_str} (added protected_at={protected_at})"
                        log_lines.append(msg)
                        print(msg)
                    except Exception as exc:
                        msg = f"FAIL {number_str} to update task.json: {exc}"
                        log_lines.append(msg)
                        print(msg)
                protected_new_count += 1
            else:
                msg = f"WARNING {number_str}: remaining but has no task.json"
                log_lines.append(msg)
                print(msg)
                missing_count += 1

    log_lines.extend(
        [
            "",
            "=== Summary ===",
            f"deleted/cleared: {deleted_count}",
            f"newly protected (stamped): {protected_new_count}",
            f"already protected (skipped): {already_protected_count}",
            f"missing/warnings: {missing_count}",
        ]
    )

    if not args.dry_run:
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        log_path.write_text("\n".join(log_lines) + "\n", encoding="utf-8")
        print(f"\nSaved log to: {log_path}")
    else:
        print("\n--- Dry Run Completed (no files written or deleted) ---")


if __name__ == "__main__":
    main()
